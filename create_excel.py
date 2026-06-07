#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Test Cases'

# Define styles
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Headers
headers = ['Test Case ID', 'Modul Uji', 'Test Type', 'Nama Test Case', 'Prekondisi', 'Langkah Pengujian', 'Data Test', 'Ekspektasi']
ws.append(headers)

# Format header row
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

# Set column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 35
ws.column_dimensions['G'].width = 20
ws.column_dimensions['H'].width = 25

# Test case data based on user input
test_cases = [
    (
        'TC01', 
        'Anggota - Aktivasi Akun', 
        'Positif', 
        'Aktivasi akun dengan OTP valid', 
        'Aplikasi terinstall dan berjalan; User sudah terdaftar; OTP screen ditampilkan', 
        'Setup: Buka aplikasi, navigasi ke screen Aktivasi Akun\nExercise: Masukkan OTP valid, tap Verifikasi\nVerify: Server memvalidasi OTP, akun diaktifkan', 
        'OTP: 123456 (benar)', 
        'Akun berhasil diaktivasi; Redirect ke halaman login / dashboard; Tampil pesan sukses'
    ),
    (
        'TC02', 
        'Anggota - Aktivasi Akun', 
        'Negatif', 
        'Aktivasi akun dengan OTP salah', 
        'Aplikasi terinstall; User sudah terdaftar; OTP screen ditampilkan', 
        'Setup: Buka aplikasi, navigasi ke screen Aktivasi Akun\nExercise: Masukkan OTP salah, tap Verifikasi\nVerify: Server menolak verifikasi', 
        'OTP: 999999 (salah)', 
        'Sistem menampilkan pesan kesalahan OTP; User tetap di screen OTP; Akun tidak aktif'
    ),
    (
        'TC03', 
        'Anggota - Login', 
        'Positif', 
        'Login dengan data valid', 
        'Aplikasi terinstall; Akun terdaftar dan aktif; Internet stabil', 
        'Setup: Buka aplikasi, navigasi ke Login Screen\nExercise: Masukkan NPA dan password benar, tap Login\nVerify: Token disimpan, navigasi ke dashboard', 
        'NPA: 12345678\nPassword: password123', 
        'Pengguna masuk ke dashboard; Sesi login aktif; Token tersimpan di secure storage'
    ),
    (
        'TC04', 
        'Anggota - Login', 
        'Negatif', 
        'Login dengan password salah', 
        'Aplikasi terinstall; Akun terdaftar; Internet stabil', 
        'Setup: Buka aplikasi, navigasi ke Login Screen\nExercise: Masukkan NPA benar dan password salah, tap Login\nVerify: Validasi server gagal', 
        'NPA: 12345678\nPassword: wrongpassword', 
        'Sistem menampilkan pesan error password salah; Sesi login gagal; User tetap di login screen'
    ),
    (
        'TC05', 
        'Anggota - Login', 
        'Negatif', 
        'Login dengan akun tidak terdaftar', 
        'Aplikasi terinstall; Akun belum terdaftar di sistem', 
        'Setup: Buka aplikasi, navigasi ke Login Screen\nExercise: Masukkan NPA tidak terdaftar dan password, tap Login\nVerify: Validasi server gagal', 
        'NPA: 99999999 (tidak terdaftar)\nPassword: password123', 
        'Sistem menampilkan pesan akun tidak ditemukan; User tetap di login screen'
    ),
    (
        'TC06', 
        'Anggota - Invoice', 
        'Positif', 
        'Melihat invoice pembayaran terakhir', 
        'User login sebagai Anggota; Transaksi pembayaran tersedia; Internet stabil', 
        'Setup: Dari dashboard Anggota, navigasi ke menu Invoice\nExercise: Pilih opsi melihat invoice pembayaran terakhir\nVerify: Tampilkan detail invoice dari server', 
        'Membuka menu invoice', 
        'Invoice terakhir ditampilkan; Detail tanggal, nominal, dan status pembayaran sesuai'
    ),
    (
        'TC07', 
        'Anggota - Invoice', 
        'Positif', 
        'Mengunduh invoice pembayaran', 
        'User login sebagai Anggota; Screen detail invoice aktif; Izin penyimpanan diberikan', 
        'Setup: Buka detail invoice pembayaran\nExercise: Klik tombol unduh (download)\nVerify: Download file PDF invoice', 
        'Klik tombol unduh', 
        'Invoice berhasil terunduh dalam format PDF; File tersimpan di local storage'
    ),
    (
        'TC08', 
        'Anggota - Riwayat Pembayaran', 
        'Positif', 
        'Melihat riwayat pembayaran', 
        'User login sebagai Anggota; Riwayat transaksi tersedia di database', 
        'Setup: Dari dashboard Anggota\nExercise: Membuka menu riwayat pembayaran\nVerify: Load daftar transaksi dari server', 
        'Membuka menu riwayat', 
        'Riwayat pembayaran ditampilkan; Menampilkan daftar transaksi yang pernah dilakukan'
    ),
    (
        'TC09', 
        'Anggota - Riwayat Pembayaran', 
        'Positif', 
        'Riwayat kosong', 
        'User login sebagai Anggota; Belum pernah melakukan transaksi pembayaran', 
        'Setup: Dari dashboard Anggota\nExercise: Membuka menu riwayat pembayaran\nVerify: Load data transaksi kosong', 
        'Tidak ada transaksi', 
        'Sistem menampilkan pesan tidak ada data; Tampilan kosong bersih tanpa error'
    ),
    (
        'TC10', 
        'Anggota - Update Profil', 
        'Positif', 
        'Mengubah nama profil', 
        'User login sebagai Anggota; Screen edit profil aktif; Internet stabil', 
        'Setup: Buka menu Profil, klik Edit Profil\nExercise: Input nama baru, klik Simpan\nVerify: Kirim update data ke server', 
        'Nama Baru: Ahmad Fauzi', 
        'Data berhasil diperbarui; Nama profil berubah di layar dan server; Pesan sukses muncul'
    ),
    (
        'TC11', 
        'Anggota - Update Profil', 
        'Positif', 
        'Mengubah nomor telepon', 
        'User login sebagai Anggota; Screen edit profil aktif; Internet stabil', 
        'Setup: Buka menu Profil, klik Edit Profil\nExercise: Input nomor telepon baru, klik Simpan\nVerify: Kirim update data ke server', 
        'No Telepon Baru: 081298765432', 
        'Data berhasil diperbarui; Nomor telepon profil terupdate di server dan UI'
    ),
    (
        'TC12', 
        'Anggota - Update Profil', 
        'Positif', 
        'Mengubah foto profil', 
        'User login sebagai Anggota; Screen edit profil aktif; Izin galeri aktif; File foto valid', 
        'Setup: Buka menu Profil, klik Edit Foto\nExercise: Pilih file foto valid, klik Simpan/Upload\nVerify: Upload foto ke server', 
        'File foto: new_avatar.jpg (valid)', 
        'Foto profil berhasil diperbarui; Foto terupdate di UI; Pesan sukses muncul'
    ),
    (
        'TC13', 
        'Anggota - Logout', 
        'Positif', 
        'Keluar dari aplikasi', 
        'User login sebagai Anggota; Berada di menu profil/pengaturan', 
        'Setup: Buka menu profil/pengaturan\nExercise: Klik tombol Logout dan konfirmasi\nVerify: Hapus token, session berakhir, navigasi ke login', 
        'Klik Logout', 
        'Kembali ke halaman login; Token dan data sesi dihapus; Tidak bisa kembali ke dashboard menggunakan tombol back'
    ),
    (
        'TC14', 
        'Bendahara PJ - Login', 
        'Positif', 
        'Login dengan data valid', 
        'Aplikasi terinstall; Akun Bendahara PJ terdaftar dan aktif; Internet stabil', 
        'Setup: Buka aplikasi, navigasi ke Login Screen\nExercise: Masukkan Username dan Password PJ benar, tap Login\nVerify: Cek role, simpan token, masuk dashboard PJ', 
        'Username: bendaharapj\nPassword: password123', 
        'Masuk ke dashboard PJ; Tampilan modul dan menu khusus Bendahara PJ terbuka'
    ),
    (
        'TC15', 
        'Bendahara PJ - Data Anggota', 
        'Positif', 
        'Melihat daftar anggota', 
        'User login sebagai Bendahara PJ; Data anggota jamaah tersedia di server', 
        'Setup: Dari dashboard PJ\nExercise: Membuka menu anggota\nVerify: Load daftar anggota yang berada di bawah PJ dari server', 
        'Membuka menu anggota', 
        'Daftar anggota ditampilkan; Informasi ringkas seperti NPA, nama, dan status iuran terlihat'
    ),
    (
        'TC16', 
        'Bendahara PJ - Data Anggota', 
        'Positif', 
        'Memilih anggota', 
        'User di list anggota Bendahara PJ; Data anggota tersedia', 
        'Setup: Buka daftar anggota\nExercise: Klik salah satu anggota\nVerify: Navigasi ke detail anggota, load detail', 
        'Klik salah satu anggota', 
        'Detail anggota ditampilkan; Menampilkan riwayat pembayaran, data profil lengkap, dan status iuran'
    ),
    (
        'TC17', 
        'Bendahara PJ - Kartu Iuran', 
        'Positif', 
        'Melihat kartu iuran anggota', 
        'User di screen detail anggota', 
        'Setup: Buka detail anggota\nExercise: Klik tombol/tab Kartu Iuran\nVerify: Aplikasi memuat tabel kartu iuran tahunan anggota tersebut', 
        'Klik kartu iuran', 
        'Data iuran anggota ditampilkan; Status pembayaran per bulan (lunas/belum) terlihat jelas'
    ),
    (
        'TC18', 
        'Bendahara PJ - Pembayaran Iuran', 
        'Positif', 
        'Memilih bulan pembayaran', 
        'User di screen kartu iuran anggota; Mode pembayaran aktif', 
        'Setup: Masuk menu Pembayaran Iuran\nExercise: Memilih bulan yang belum dibayar (misal: Januari)\nVerify: Bulan terpilih disorot, sistem menyiapkan nominal iuran', 
        'Memilih bulan yang belum dibayar', 
        'Data bulan terpilih; Tombol bayar menjadi aktif dengan nominal yang sesuai'
    ),
    (
        'TC19', 
        'Bendahara PJ - Pembayaran Iuran', 
        'Positif', 
        'Pembayaran bulan berurutan', 
        'User di screen Pembayaran Iuran; Bulan yang dipilih berurutan dengan pembayaran terakhir', 
        'Setup: Pilih bulan berikutnya (misal: Februari setelah Januari lunas)\nExercise: Membayar bulan berikutnya, tap Proses Bayar\nVerify: Sistem memvalidasi urutan bulan, memproses pembayaran', 
        'Membayar bulan berikutnya', 
        'Pembayaran berhasil diproses; Status bulan berubah menjadi lunas; Saldo terupdate'
    ),
    (
        'TC20', 
        'Bendahara PJ - Pembayaran Iuran', 
        'Negatif', 
        'Pembayaran bulan tidak berurutan', 
        'User di screen Pembayaran Iuran; Ada bulan sebelumnya yang belum lunas', 
        'Setup: Pilih bulan melompati bulan belum lunas (misal: Maret padahal Februari belum lunas)\nExercise: Klik Proses Bayar\nVerify: Sistem mengecek urutan bulan, menolak transaksi', 
        'Melompati bulan sebelumnya', 
        'Sistem menolak transaksi dan menampilkan error; Pesan: "Pembayaran harus berurutan, selesaikan bulan sebelumnya"'
    ),
    (
        'TC21', 
        'Bendahara PJ - Pembayaran Iuran', 
        'Positif', 
        'Pembayaran saat online', 
        'User terkoneksi internet; Form pembayaran iuran siap', 
        'Setup: Koneksi internet aktif\nExercise: Proses pembayaran iuran, tap Simpan\nVerify: Kirim data langsung ke server API', 
        'Internet tersedia', 
        'Status pembayaran menjadi lunas di server; Data ter-update secara real-time di database server'
    ),
    (
        'TC22', 
        'Bendahara PJ - Pembayaran Iuran', 
        'Positif', 
        'Pembuatan invoice otomatis', 
        'Pembayaran iuran sukses dilakukan', 
        'Setup: Transaksi sukses tersimpan\nVerify: Sistem memicu generator invoice\nVerify: Invoice PDF baru ditambahkan ke log transaksi', 
        'Pembayaran berhasil', 
        'Invoice otomatis dibuat; Muncul notifikasi invoice siap dan tombol untuk cetak/kirim'
    ),
    (
        'TC23', 
        'Bendahara PJ - Invoice', 
        'Positif', 
        'Mengirim invoice ke WhatsApp', 
        'Invoice telah dibuat; WhatsApp terinstall di device', 
        'Setup: Buka detail invoice / transaksi\nExercise: Klik Kirim WhatsApp\nVerify: Aplikasi membuka share sheet ke WA dengan attachment/link invoice', 
        'Klik Kirim WhatsApp', 
        'Invoice berhasil dikirim; WhatsApp terbuka dengan kontak/nomor anggota terpilih'
    ),
    (
        'TC24', 
        'Bendahara PJ - Pembayaran Iuran', 
        'Positif', 
        'Pembayaran saat offline', 
        'Koneksi internet terputus; Form pembayaran iuran siap', 
        'Setup: Matikan internet\nExercise: Lakukan pembayaran iuran, tap Simpan\nVerify: Sistem mendeteksi offline, menyimpan transaksi ke Hive lokal', 
        'Internet tidak tersedia', 
        'Data tersimpan di Hive/local storage; Transaksi ditandai status "pending sync"; Selesai tanpa crash'
    ),
    (
        'TC25', 
        'Bendahara PJ - Sinkronisasi Data', 
        'Positif', 
        'Sinkronisasi otomatis setelah online', 
        'Ada transaksi berstatus "pending sync" di Hive; Internet kembali terhubung', 
        'Setup: Aktifkan kembali internet\nVerify: Aplikasi mendeteksi koneksi dan melakukan background auto-sync\nVerify: Mengirim data ke server API', 
        'Koneksi internet tersedia kembali', 
        'Data berhasil dikirim ke server; Status transaksi lokal berubah menjadi lunas/synced; Banner offline hilang'
    ),
    (
        'TC26', 
        'Bendahara PJ - Log Transaksi', 
        'Positif', 
        'Melihat log transaksi', 
        'User login Bendahara PJ; Transaksi kas PJ sudah pernah dilakukan', 
        'Setup: Dari dashboard PJ\nExercise: Membuka menu log transaksi\nVerify: Aplikasi load daftar transaksi historis dari local/server', 
        'Membuka menu log transaksi', 
        'Histori transaksi ditampilkan; Menampilkan daftar masuk dan keluar beserta status sync'
    ),
    (
        'TC27', 
        'Bendahara PJ - Log Transaksi', 
        'Positif', 
        'Log transaksi kosong', 
        'User login Bendahara PJ; Belum ada transaksi sama sekali', 
        'Setup: Dari dashboard PJ\nExercise: Membuka menu log transaksi\nVerify: Load data transaksi kosong', 
        'Tidak ada transaksi', 
        'Sistem menampilkan data kosong; Pesan "Belum ada transaksi" ditampilkan'
    ),
    (
        'TC28', 
        'Bendahara PJ - Laporan', 
        'Positif', 
        'Melihat laporan pembayaran', 
        'User login Bendahara PJ; Data transaksi bulanan tersedia', 
        'Setup: Buka menu Laporan\nExercise: Memilih bulan dan tahun laporan, tap Cari/Tampilkan\nVerify: Aplikasi memuat summary rekapitulasi periode tersebut', 
        'Memilih bulan dan tahun', 
        'Rekap pembayaran ditampilkan; Menampilkan total nominal lunas, tunggakan, dan grafik ringkasan'
    ),
    (
        'TC29', 
        'Bendahara PJ - Laporan', 
        'Positif', 
        'Filter laporan berdasarkan bulan', 
        'User berada di menu Laporan', 
        'Setup: Buka menu Laporan\nExercise: Memilih bulan tertentu (misal: Juni)\nVerify: Aplikasi menyaring list laporan transaksi hanya untuk bulan Juni', 
        'Memilih bulan tertentu', 
        'Data sesuai filter ditampilkan; Hanya transaksi bulan terpilih yang muncul'
    ),
    (
        'TC30', 
        'Bendahara PJ - Laporan', 
        'Positif', 
        'Filter laporan berdasarkan tahun', 
        'User berada di menu Laporan', 
        'Setup: Buka menu Laporan\nExercise: Memilih tahun tertentu (misal: 2026)\nVerify: Aplikasi menyaring list laporan transaksi hanya untuk tahun 2026', 
        'Memilih tahun tertentu', 
        'Data sesuai filter ditampilkan; Hanya transaksi tahun terpilih yang muncul'
    ),
    (
        'TC31', 
        'Bendahara PJ - Laporan', 
        'Positif', 
        'Laporan tanpa data', 
        'User berada di menu Laporan; Periode terpilih tidak memiliki transaksi', 
        'Setup: Buka menu Laporan\nExercise: Pilih bulan/tahun tanpa aktivitas transaksi\nVerify: Tampilkan pesan data kosong', 
        'Periode tidak memiliki transaksi', 
        'Sistem menampilkan pesan tidak ada data; Grafik kosong'
    ),
    (
        'TC32', 
        'Bendahara PJ - Cetak Laporan', 
        'Positif', 
        'Mengunduh laporan pembayaran', 
        'User berada di screen Laporan PJ; Izin penyimpanan diberikan', 
        'Setup: Buka menu Laporan\nExercise: Klik Cetak Laporan\nVerify: Generate excel file (.xlsx) laporan', 
        'Klik Cetak Laporan', 
        'File .xlsx berhasil terunduh; Struktur tabel laporan rapi dan data valid'
    ),
    (
        'TC33', 
        'Bendahara PJ - Logout', 
        'Positif', 
        'Keluar dari aplikasi', 
        'User login Bendahara PJ; Berada di menu utama/profil', 
        'Setup: Buka menu profil/pengaturan\nExercise: Klik Logout\nVerify: Hapus token, navigasi kembali ke login', 
        'Klik Logout', 
        'Kembali ke halaman login; Sesi dan token dihapus'
    ),
    (
        'TC34', 
        'Bendahara PC - Login', 
        'Positif', 
        'Login dengan data valid', 
        'Aplikasi terinstall; Akun Bendahara PC terdaftar dan aktif; Internet stabil', 
        'Setup: Buka login screen\nExercise: Masukkan Username dan Password PC benar, tap Login\nVerify: Validasi role PC, simpan token, masuk dashboard PC', 
        'Username: bendaharapc\nPassword: password123', 
        'Masuk ke dashboard PC; Tampilan modul dan menu khusus Bendahara PC terbuka'
    ),
    (
        'TC35', 
        'Bendahara PC - Laporan', 
        'Positif', 
        'Melihat laporan pembayaran', 
        'User login Bendahara PC; Data transaksi cabang (PC) tersedia', 
        'Setup: Buka menu Laporan PC\nExercise: Memilih bulan dan tahun laporan, klik Tampilkan\nVerify: Load summary rekapitulasi seluruh cabang/jamaah', 
        'Memilih bulan dan tahun', 
        'Rekap pembayaran ditampilkan; Menampilkan total nominal lunas dan grafik ringkasan cabang'
    ),
    (
        'TC36', 
        'Bendahara PC - Laporan', 
        'Positif', 
        'Filter laporan berdasarkan bulan', 
        'User berada di menu Laporan PC', 
        'Setup: Buka menu Laporan PC\nExercise: Memilih bulan tertentu\nVerify: Aplikasi menyaring laporan transaksi PC berdasarkan bulan', 
        'Memilih bulan tertentu', 
        'Data sesuai filter ditampilkan; Hanya laporan bulan terpilih yang muncul'
    ),
    (
        'TC37', 
        'Bendahara PC - Laporan', 
        'Positif', 
        'Filter laporan berdasarkan tahun', 
        'User berada di menu Laporan PC', 
        'Setup: Buka menu Laporan PC\nExercise: Memilih tahun tertentu\nVerify: Aplikasi menyaring laporan transaksi PC berdasarkan tahun', 
        'Memilih tahun tertentu', 
        'Data sesuai filter ditampilkan; Hanya laporan tahun terpilih yang muncul'
    ),
    (
        'TC38', 
        'Bendahara PC - Laporan', 
        'Positif', 
        'Laporan tanpa data', 
        'User berada di menu Laporan PC; Periode terpilih tidak memiliki transaksi', 
        'Setup: Buka menu Laporan PC\nExercise: Pilih bulan/tahun tanpa aktivitas transaksi\nVerify: Tampilkan pesan data kosong', 
        'Periode tidak memiliki transaksi', 
        'Sistem menampilkan pesan tidak ada data; Grafik kosong'
    ),
    (
        'TC39', 
        'Bendahara PC - Rekapitulasi', 
        'Positif', 
        'Melihat rekap pembayaran seluruh jamaah', 
        'User login Bendahara PC; Data rekap jamaah tersedia', 
        'Setup: Dari dashboard PC\nExercise: Membuka menu rekapitulasi\nVerify: Load rekapitulasi pembayaran dari semua cabang/PJ', 
        'Membuka menu rekap', 
        'Rekap pembayaran ditampilkan; Menampilkan total lunas, total tertunggak per unit PJ'
    ),
    (
        'TC40', 
        'Bendahara PC - Rekapitulasi', 
        'Positif', 
        'Melihat detail rekap pembayaran', 
        'User di screen Rekapitulasi Bendahara PC', 
        'Setup: Buka menu rekapitulasi\nExercise: Memilih salah satu data rekap PJ/Jamaah\nVerify: Navigasi dan load daftar detail transaksi dari PJ terpilih', 
        'Memilih salah satu data rekap', 
        'Detail transaksi ditampilkan; Menampilkan rincian pembayaran anggota-anggota di bawah PJ tersebut'
    ),
    (
        'TC41', 
        'Bendahara PC - Cetak Laporan', 
        'Positif', 
        'Mengunduh laporan rekapitulasi', 
        'User berada di menu Rekapitulasi PC; Izin penyimpanan diberikan', 
        'Setup: Buka menu rekapitulasi/laporan\nExercise: Klik Cetak Laporan\nVerify: Generate excel file (.xlsx) rekapitulasi', 
        'Klik Cetak Laporan', 
        'File .xlsx berhasil terunduh; Tabel rekapitulasi seluruh cabang/jamaah tersimpan'
    ),
    (
        'TC42', 
        'Bendahara PC - Logout', 
        'Positif', 
        'Keluar dari aplikasi', 
        'User login Bendahara PC; Berada di menu profil/pengaturan', 
        'Setup: Buka profil/pengaturan\nExercise: Klik Logout\nVerify: Hapus token, navigasi kembali to login', 
        'Klik Logout', 
        'Kembali ke halaman login; Sesi dan token dihapus'
    )
]

# Add test cases to sheet
for i, tc in enumerate(test_cases, 2):
    ws.append(tc)
    for j, cell in enumerate(ws[i]):
        cell.border = border
        if j == 2:  # Test Type column (index 2)
            cell.alignment = center_align
        else:
            cell.alignment = left_align

# Format all cells
for row in ws.iter_rows(min_row=2, max_row=len(test_cases)+1):
    for cell in row:
        cell.font = Font(size=9)

# Freeze header row
ws.freeze_panes = 'A2'

# Save workbook
wb.save('TEST_CASES_BLACKBOX.xlsx')
print('Excel file created successfully: TEST_CASES_BLACKBOX.xlsx')
print(f'Total test cases: {len(test_cases)}')

# Generate TEST_CASES_BLACKBOX.md (Detailed format)
with open('TEST_CASES_BLACKBOX.md', 'w', encoding='utf-8') as f:
    f.write("# Test Case Blackbox Testing - PERSIS App\n\n")
    
    current_module = ""
    for tc in test_cases:
        tc_id, modul_uji, test_type, nama_test_case, prekondisi, langkah_pengujian, data_test, ekspektasi = tc
        
        main_module = modul_uji.split(" - ")[0]
        if main_module != current_module:
            current_module = main_module
            f.write(f"## FEATURE: {current_module.upper()}\n\n")
            
        f.write(f"### {tc_id}: {nama_test_case}\n")
        f.write("| Aspek | Detail |\n")
        f.write("|-------|--------|\n")
        f.write(f"| **Modul Uji** | {modul_uji} |\n")
        f.write(f"| **Test Type** | {test_type} |\n")
        f.write(f"| **Nama Test Case** | {nama_test_case} |\n")
        
        pre_formatted = prekondisi.replace('\n', '<br>')
        step_formatted = langkah_pengujian.replace('\n', '<br>')
        data_formatted = data_test.replace('\n', '<br>')
        exp_formatted = ekspektasi.replace('\n', '<br>')
        
        f.write(f"| **Prekondisi** | {pre_formatted} |\n")
        f.write(f"| **Langkah Pengujian** | {step_formatted} |\n")
        f.write(f"| **Data Test** | {data_formatted} |\n")
        f.write(f"| **Ekspektasi** | {exp_formatted} |\n\n")

print('Markdown file created successfully: TEST_CASES_BLACKBOX.md')

# Generate TEST_CASES_BLACKBOX_FORMATTED.md (Table format)
with open('TEST_CASES_BLACKBOX_FORMATTED.md', 'w', encoding='utf-8') as f:
    f.write("# Test Case Blackbox Testing - PERSIS App\n\n")
    
    current_module = ""
    for tc in test_cases:
        tc_id, modul_uji, test_type, nama_test_case, prekondisi, langkah_pengujian, data_test, ekspektasi = tc
        
        main_module = modul_uji.split(" - ")[0]
        if main_module != current_module:
            current_module = main_module
            f.write(f"## {current_module.upper()} MODULE\n\n")
            f.write("| Test Case ID | Modul Uji | Test Type | Nama Test Case | Prekondisi | Langkah Pengujian | Data Test | Ekspektasi |\n")
            f.write("|---|---|---|---|---|---|---|---|\n")
            
        pre_formatted = prekondisi.replace('\n', '<br>')
        step_formatted = langkah_pengujian.replace('\n', '<br>')
        data_formatted = data_test.replace('\n', '<br>')
        exp_formatted = ekspektasi.replace('\n', '<br>')
        
        f.write(f"| {tc_id} | {modul_uji} | {test_type} | {nama_test_case} | {pre_formatted} | {step_formatted} | {data_formatted} | {exp_formatted} |\n")

    f.write("\n---\n\n## RINGKASAN STATISTIK TEST CASE\n\n")
    f.write("| Feature | Total | Positif | Negatif |\n")
    f.write("|---------|-------|---------|---------|\n")
    f.write("| Anggota | 13 | 10 | 3 |\n")
    f.write("| Bendahara PJ | 20 | 19 | 1 |\n")
    f.write("| Bendahara PC | 9 | 9 | 0 |\n")
    f.write("| **TOTAL** | **42** | **38** | **4** |\n")

print('Formatted Markdown file created successfully: TEST_CASES_BLACKBOX_FORMATTED.md')
